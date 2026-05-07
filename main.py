from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
import uvicorn, os, uuid, re, httpx, asyncio, io, json, tempfile
from datetime import datetime

app = FastAPI(title="Enriquece")
sessions = {}

HUNTER_KEY = os.getenv("HUNTER_API_KEY", "")
APOLLO_KEY  = os.getenv("APOLLO_API_KEY", "")

def clean_cnpj(raw):
    c = re.sub(r"\D", "", str(raw).strip())
    return c.zfill(14) if len(c) < 14 else c

def validate_cnpj(cnpj):
    c = clean_cnpj(cnpj)
    if len(c) != 14 or len(set(c)) == 1:
        return False
    w1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    s1 = sum(int(c[i]) * w1[i] for i in range(12))
    d1 = 0 if (11 - s1 % 11) >= 10 else (11 - s1 % 11)
    w2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    s2 = sum(int(c[i]) * w2[i] for i in range(13))
    d2 = 0 if (11 - s2 % 11) >= 10 else (11 - s2 % 11)
    return int(c[12]) == d1 and int(c[13]) == d2

def format_cnpj(c):
    c = clean_cnpj(c)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else c

def not_empty(v):
    return bool(v) and str(v).strip() not in ("", "None", "não encontrado", "—")

def score_cargo(cargo):
    c = (cargo or "").lower()
    for k in ["parcerias","comercial","financeiro","cfo","marketplace","sellers","payments","pagamento","revenue","growth"]:
        if k in c: return "Alta"
    for k in ["produto","operações","gerente"]:
        if k in c: return "Média"
    return "Baixa"

def motivo_cargo(cargo):
    c = (cargo or "").lower()
    if "parcerias" in c: return "Decisor direto de parcerias e integrações"
    if "comercial" in c or "sales" in c: return "Responsável por vendas e receita"
    if "financeiro" in c or "cfo" in c: return "Decisor sobre crédito e antecipação"
    if "marketplace" in c or "sellers" in c: return "Gestão de sellers — cliente CredSeller"
    return "Contato relevante para abordagem comercial"

async def fetch_brasilapi(cnpj_clean, client):
    try:
        r = await client.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_clean}", timeout=10)
        if r.status_code == 200:
            return r.json(), "BrasilAPI"
    except: pass
    return None, None

async def fetch_receitaws(cnpj_clean, client):
    try:
        r = await client.get(f"https://www.receitaws.com.br/v1/cnpj/{cnpj_clean}", timeout=10)
        if r.status_code == 200:
            return r.json(), "ReceitaWS"
    except: pass
    return None, None

async def fetch_cnpjws(cnpj_clean, client):
    try:
        r = await client.get(f"https://publica.cnpj.ws/cnpj/{cnpj_clean}", timeout=10)
        if r.status_code == 200:
            return r.json(), "CNPJ.ws"
    except: pass
    return None, None

def parse_cadastral(data, source):
    if not data: return {}
    r = {}
    if source == "BrasilAPI":
        r["razao_social"]   = data.get("razao_social","")
        r["nome_fantasia"]  = data.get("nome_fantasia","")
        r["situacao"]       = data.get("descricao_situacao_cadastral","")
        r["cnae_principal"] = data.get("cnae_fiscal_descricao","")
        r["natureza_jur"]   = data.get("descricao_natureza_juridica","")
        r["porte"]          = data.get("porte","")
        r["cidade"]         = data.get("municipio","")
        r["estado"]         = data.get("uf","")
        r["cep"]            = data.get("cep","")
        r["telefone_cad"]   = data.get("ddd_telefone_1","") or data.get("ddd_telefone_2","")
        r["email_cad"]      = data.get("email","")
        r["endereco"]       = ", ".join(filter(None,[data.get("logradouro",""),data.get("numero",""),data.get("bairro","")]))
    elif source == "ReceitaWS":
        r["razao_social"]   = data.get("nome","")
        r["nome_fantasia"]  = data.get("fantasia","")
        r["situacao"]       = data.get("situacao","")
        r["cnae_principal"] = (data.get("atividade_principal") or [{}])[0].get("text","")
        r["natureza_jur"]   = data.get("natureza_juridica","")
        r["porte"]          = data.get("porte","")
        r["cidade"]         = data.get("municipio","")
        r["estado"]         = data.get("uf","")
        r["cep"]            = data.get("cep","")
        r["telefone_cad"]   = data.get("telefone","")
        r["email_cad"]      = data.get("email","")
        r["endereco"]       = data.get("logradouro","")
    elif source == "CNPJ.ws":
        est = data.get("estabelecimento") or {}
        r["razao_social"]   = data.get("razao_social","")
        r["nome_fantasia"]  = est.get("nome_fantasia","")
        r["situacao"]       = est.get("situacao_cadastral","")
        r["cnae_principal"] = (est.get("atividade_principal") or {}).get("descricao","")
        r["natureza_jur"]   = (data.get("natureza_juridica") or {}).get("descricao","")
        r["porte"]          = (data.get("porte") or {}).get("descricao","")
        r["cidade"]         = (est.get("cidade") or {}).get("nome","")
        r["estado"]         = (est.get("estado") or {}).get("sigla","")
        r["cep"]            = est.get("cep","")
        tel = est.get("ddd1","") + est.get("telefone1","") if est.get("telefone1") else ""
        r["telefone_cad"]   = tel
        r["email_cad"]      = est.get("email","")
        r["endereco"]       = ", ".join(filter(None,[est.get("logradouro",""),est.get("numero",""),est.get("bairro","")]))
    return {k:v for k,v in r.items() if v}

async def fetch_hunter(domain, client):
    if not HUNTER_KEY or not domain: return {}
    try:
        r = await client.get("https://api.hunter.io/v2/domain-search",
            params={"domain": domain, "api_key": HUNTER_KEY, "limit": 5}, timeout=10)
        if r.status_code == 200:
            return r.json().get("data", {})
    except: pass
    return {}

async def fetch_apollo(company, domain, client):
    if not APOLLO_KEY: return {}
    try:
        payload = {
            "api_key": APOLLO_KEY,
            "q_organization_name": company,
            "organization_domains": [domain] if domain else [],
            "person_titles": ["Diretor de Parcerias","Head de Parcerias","Diretor Comercial",
                "Head Comercial","CFO","Diretor Financeiro","Head Financeiro",
                "Diretor de Marketplace","Head de Marketplace","Gerente Comercial"],
            "page": 1, "per_page": 5,
        }
        r = await client.post("https://api.apollo.io/v1/mixed_people/search", json=payload, timeout=12)
        if r.status_code == 200:
            return r.json()
    except: pass
    return {}

@app.get("/", response_class=HTMLResponse)
def index():
    try:
        with open("index.html","r",encoding="utf-8") as f: return f.read()
    except: return "<h1>Enriquece online!</h1>"

@app.get("/health")
def health():
    return {"status":"ok","hunter":"ok" if HUNTER_KEY else "sem chave","apollo":"ok" if APOLLO_KEY else "sem chave"}

@app.post("/api/cnpj/upload")
async def upload(file: UploadFile = File(...)):
    content = await file.read()
    rows, headers = [], []
    try:
        if file.filename.lower().endswith(".csv"):
            import csv
            decoded = content.decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
            headers = list(rows[0].keys()) if rows else []
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value) if c.value is not None else f"col{i}" for i,c in enumerate(ws[1])]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
    except Exception as e:
        raise HTTPException(400, f"Erro: {e}")

    col = None
    for h in headers:
        sample = [str(r.get(h,"") or "") for r in rows[:15]]
        if sum(1 for v in sample if 11 <= len(clean_cnpj(v)) <= 14) >= 1:
            col = h; break
    if not col: col = headers[0] if headers else "CNPJ"

    seen, valid, dups, invalid = set(), [], [], []
    for i, row in enumerate(rows):
        raw = str(row.get(col,"") or "")
        if not raw or raw=="None": continue
        c = clean_cnpj(raw)
        if not validate_cnpj(c):
            invalid.append({"raw":raw,"row":i+2,"reason":"CNPJ inválido"})
        elif c in seen:
            dups.append({"raw":raw,"row":i+2,"reason":"Duplicado"})
        else:
            seen.add(c)
            valid.append({"cnpj":format_cnpj(c),"clean":c})

    sid = str(uuid.uuid4())
    sessions[sid] = {"status":"validated","filename":file.filename,
                     "valid":valid,"invalid_rows":invalid+dups,"results":[],"total":len(rows)}
    return {"session_id":sid,"filename":file.filename,"total":len(rows),
            "valid":len(valid),"duplicates":len(dups),"invalid":len(invalid),
            "lots":max(1,-(-len(valid)//500)),
            "preview":[{"cnpj":r["cnpj"],"status":"válido"} for r in valid[:10]]}

@app.post("/api/enrich/start/{sid}")
def start(sid: str, lot_index: int=0):
    if sid not in sessions: raise HTTPException(404)
    sessions[sid]["status"] = "processing"
    return {"ok":True}

@app.get("/api/enrich/stream/{sid}")
async def stream(sid: str, lot_index: int=0):
    async def gen():
        if sid not in sessions:
            yield "data: "+json.dumps({"error":"Sessão não encontrada"})+"\n\n"; return
        items = sessions[sid]["valid"]
        results = []
        async with httpx.AsyncClient(headers={"User-Agent":"Enriquece/1.0"},follow_redirects=True,timeout=12) as client:
            for i, item in enumerate(items):
                r = {"cnpj":item["cnpj"],"clean":item["clean"],
                     "razao_social":"não encontrado","nome_fantasia":"não encontrado",
                     "situacao":"não encontrado","cnae_principal":"não encontrado",
                     "natureza_jur":"não encontrado","porte":"não encontrado",
                     "cidade":"não encontrado","estado":"não encontrado",
                     "cep":"não encontrado","endereco":"não encontrado",
                     "telefone_cad":"não encontrado","telefone_extra":"não encontrado",
                     "email_cad":"não encontrado","email_comercial":"não encontrado",
                     "email_financeiro":"não encontrado",
                     "site":"não encontrado","linkedin_empresa":"não encontrado",
                     "score_geral":"baixa","status":"sem dados","decisores":[],"fontes":[]}

                # 1. Dados cadastrais — BrasilAPI > ReceitaWS > CNPJ.ws
                data, src = await fetch_brasilapi(item["clean"], client)
                if not data: data, src = await fetch_receitaws(item["clean"], client)
                if not data: data, src = await fetch_cnpjws(item["clean"], client)
                if data and src:
                    for k,v in parse_cadastral(data,src).items():
                        if not_empty(v): r[k] = v
                    r["fontes"].append({"dado":"Dados cadastrais","fonte":src,"confianca":"alta","data":datetime.now().strftime("%d/%m/%Y")})

                company = r["razao_social"] if not_empty(r["razao_social"]) else ""
                fantasia = r["nome_fantasia"] if not_empty(r["nome_fantasia"]) else company
                domain = ""

                # 2. Hunter.io — e-mails
                if HUNTER_KEY and fantasia:
                    slug = re.sub(r"[^a-z0-9]","",fantasia.lower()[:25])
                    for ext in [".com.br",".com"]:
                        hdata = await fetch_hunter(slug+ext, client)
                        if hdata.get("emails"):
                            domain = hdata.get("domain","")
                            for e in hdata["emails"][:4]:
                                em = e.get("value","")
                                pos = (e.get("position") or "").lower()
                                if not_empty(em):
                                    if "financ" in pos and not not_empty(r["email_financeiro"]):
                                        r["email_financeiro"] = em
                                    elif ("comerci" in pos or "sales" in pos) and not not_empty(r["email_comercial"]):
                                        r["email_comercial"] = em
                                    elif not not_empty(r["email_cad"]):
                                        r["email_cad"] = em
                            r["fontes"].append({"dado":"E-mails","fonte":"Hunter.io","confianca":"alta","data":datetime.now().strftime("%d/%m/%Y")})
                            break

                # 3. Apollo.io — decisores
                if APOLLO_KEY and company:
                    adata = await fetch_apollo(company, domain, client)
                    for p in (adata.get("people") or [])[:5]:
                        nome = p.get("name","")
                        cargo = p.get("title","")
                        if nome and cargo:
                            r["decisores"].append({
                                "nome":nome,"cargo":cargo,
                                "area":(p.get("departments") or [""])[0],
                                "email":p.get("email",""),
                                "linkedin":p.get("linkedin_url",""),
                                "phone":((p.get("phone_numbers") or [{}])[0].get("raw_number","")),
                                "prioridade":score_cargo(cargo),
                                "motivo":motivo_cargo(cargo),
                                "score":"alta" if p.get("email") else "media",
                                "fonte":"Apollo.io",
                            })
                    if r["decisores"]:
                        r["fontes"].append({"dado":"Decisores","fonte":"Apollo.io","confianca":"alta","data":datetime.now().strftime("%d/%m/%Y")})

                # Score
                pts = 0
                if not_empty(r["telefone_cad"]): pts+=2
                if not_empty(r["email_cad"]) or not_empty(r["email_comercial"]): pts+=2
                if not_empty(r["site"]): pts+=1
                if r["decisores"]: pts+=3
                if not_empty(r["razao_social"]): pts+=1
                r["score_geral"] = "alta" if pts>=6 else "media" if pts>=2 else "baixa"
                r["status"] = "completo" if pts>=5 else "parcial" if pts>=1 else "sem dados"

                results.append(r)
                sessions[sid]["results"] = results
                pct = round((i+1)/len(items)*100)
                yield "data: "+json.dumps({
                    "status":"processing","processed":i+1,"total":len(items),"pct":pct,
                    "current_cnpj":item["cnpj"],"current_company":r["razao_social"],
                    "phones_found":sum(1 for x in results if not_empty(x.get("telefone_cad"))),
                    "emails_found":sum(1 for x in results if not_empty(x.get("email_cad")) or not_empty(x.get("email_comercial"))),
                    "contacts_found":sum(1 for x in results if not_empty(x.get("email_cad"))),
                    "decisors_found":sum(1 for x in results if x.get("decisores")),
                })+"\n\n"
                await asyncio.sleep(0.15)

        sessions[sid]["status"] = "done"
        n = len(results)
        p = lambda x: round(x/n*100) if n else 0
        yield "data: "+json.dumps({
            "status":"done","total":n,
            "pct_telefone":p(sum(1 for x in results if not_empty(x.get("telefone_cad")))),
            "pct_email":p(sum(1 for x in results if not_empty(x.get("email_cad")) or not_empty(x.get("email_comercial")))),
            "pct_site":p(sum(1 for x in results if not_empty(x.get("site")))),
            "pct_decisor":p(sum(1 for x in results if x.get("decisores"))),
            "pct_alta":p(sum(1 for x in results if x.get("score_geral")=="alta")),
        })+"\n\n"

    return StreamingResponse(gen(), media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.get("/api/enrich/results/{sid}")
def results(sid: str):
    if sid not in sessions: raise HTTPException(404)
    return {"results":sessions[sid].get("results",[]),"status":sessions[sid].get("status")}

@app.post("/api/enrich/pause/{sid}")
def pause(sid: str): return {"ok":True}

@app.post("/api/enrich/resume/{sid}")
def resume(sid: str): return {"ok":True}

@app.get("/api/export/excel/{sid}")
def export(sid: str, filter: str="all"):
    if sid not in sessions: raise HTTPException(404)
    results = sessions[sid].get("results",[])
    if filter=="alta": results=[r for r in results if r.get("score_geral")=="alta"]
    elif filter=="decisores": results=[r for r in results if r.get("decisores")]
    elif filter=="pendencias": results=[r for r in results if r.get("status")!="completo"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    navy = "0D1B3E"; white="FFFFFF"; lgray="F4F6FB"
    hfill = PatternFill("solid",fgColor=navy)
    hfont = Font(bold=True,color=white,name="Arial",size=11)
    score_colors = {"alta":"E6F9F4","media":"FFF8E6","baixa":"FEF2F2"}

    # ABA 1 — Base Enriquecida
    ws1 = wb.active; ws1.title="Base Enriquecida"; ws1.sheet_view.showGridLines=False
    h1=["CNPJ","Razão Social","Nome Fantasia","Situação","CNAE","Natureza Jurídica","Porte",
        "Cidade","Estado","CEP","Endereço","Telefone","E-mail Geral","E-mail Comercial",
        "E-mail Financeiro","Site","LinkedIn","Score","Status"]
    for i,h in enumerate(h1,1):
        c=ws1.cell(row=1,column=i,value=h); c.fill=hfill; c.font=hfont
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height=30
    for idx,r in enumerate(results,2):
        bg=PatternFill("solid",fgColor=lgray if idx%2==0 else white)
        vals=[r.get("cnpj",""),r.get("razao_social",""),r.get("nome_fantasia",""),
              r.get("situacao",""),r.get("cnae_principal",""),r.get("natureza_jur",""),
              r.get("porte",""),r.get("cidade",""),r.get("estado",""),r.get("cep",""),
              r.get("endereco",""),r.get("telefone_cad",""),r.get("email_cad",""),
              r.get("email_comercial",""),r.get("email_financeiro",""),
              r.get("site",""),r.get("linkedin_empresa",""),
              (r.get("score_geral","") or "").upper(),r.get("status","")]
        for j,v in enumerate(vals,1):
            cell=ws1.cell(row=idx,column=j,value=v if v not in ("não encontrado","") else "—")
            cell.fill=bg; cell.font=Font(name="Arial",size=10)
            if j==18:
                sc=r.get("score_geral","baixa")
                cell.fill=PatternFill("solid",fgColor=score_colors.get(sc,lgray))
                cell.font=Font(bold=True,name="Arial",size=10)
    ws1.freeze_panes="A2"; ws1.auto_filter.ref=f"A1:S{len(results)+1}"
    for i,w in enumerate([20,35,25,14,38,28,14,18,8,12,38,18,30,30,30,28,28,10,12],1):
        ws1.column_dimensions[get_column_letter(i)].width=w

    # ABA 2 — Decisores
    ws2=wb.create_sheet("Decisores"); ws2.sheet_view.showGridLines=False
    h2=["CNPJ","Empresa","Nome Decisor","Cargo","Área","LinkedIn","E-mail","Telefone","Prioridade","Motivo","Score","Fonte"]
    bfill=PatternFill("solid",fgColor="1E4DB7")
    for i,h in enumerate(h2,1):
        c=ws2.cell(row=1,column=i,value=h); c.fill=bfill
        c.font=Font(bold=True,color=white,name="Arial",size=11)
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=30
    row2=2
    for r in results:
        for d in r.get("decisores",[]):
            bg=PatternFill("solid",fgColor=lgray if row2%2==0 else white)
            vals=[r.get("cnpj",""),r.get("razao_social",""),d.get("nome",""),d.get("cargo",""),
                  d.get("area",""),d.get("linkedin",""),d.get("email",""),d.get("phone",""),
                  d.get("prioridade",""),d.get("motivo",""),d.get("score","").upper(),d.get("fonte","")]
            for j,v in enumerate(vals,1):
                cell=ws2.cell(row=row2,column=j,value=v or "—")
                cell.fill=bg; cell.font=Font(name="Arial",size=10)
            row2+=1
    ws2.freeze_panes="A2"
    for i,w in enumerate([20,30,25,28,18,35,30,16,12,45,10,12],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    # ABA 3 — Resumo
    ws3=wb.create_sheet("Resumo"); ws3.sheet_view.showGridLines=False
    ws3.merge_cells("A1:B1"); c=ws3["A1"]
    c.value="ENRIQUECE — Resumo do Processamento"
    c.font=Font(bold=True,color=white,size=14,name="Arial")
    c.fill=PatternFill("solid",fgColor=navy); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws3.row_dimensions[1].height=35
    n=len(results)
    kv=[("Total processado",n),
        ("Com telefone",sum(1 for r in results if not_empty(r.get("telefone_cad")))),
        ("Com e-mail",sum(1 for r in results if not_empty(r.get("email_cad")) or not_empty(r.get("email_comercial")))),
        ("Com site",sum(1 for r in results if not_empty(r.get("site")))),
        ("Com decisor",sum(1 for r in results if r.get("decisores"))),
        ("Alta confiança",sum(1 for r in results if r.get("score_geral")=="alta")),
        ("Data",datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Gerado por","Enriquece · CredSeller")]
    for i,(k,v) in enumerate(kv,3):
        ws3.cell(row=i,column=1,value=k).font=Font(bold=True,name="Arial",size=11)
        ws3.cell(row=i,column=2,value=v).font=Font(name="Arial",size=11)
        if i%2==0:
            for col in range(1,3):
                ws3.cell(row=i,column=col).fill=PatternFill("solid",fgColor=lgray)
    ws3.column_dimensions["A"].width=25; ws3.column_dimensions["B"].width=30

    # ABA 4 — Pendências
    ws4=wb.create_sheet("Pendências"); ws4.sheet_view.showGridLines=False
    h4=["CNPJ","Empresa","Campo Pendente","Motivo","Próxima Ação","Prioridade"]
    afill=PatternFill("solid",fgColor="F59E0B")
    for i,h in enumerate(h4,1):
        c=ws4.cell(row=1,column=i,value=h); c.fill=afill
        c.font=Font(bold=True,color=navy,name="Arial",size=11)
    row4=2
    campos=[("telefone_cad","Telefone","Buscar no site ou Google Maps","Alta"),
            ("email_cad","E-mail","Buscar via Hunter.io","Alta"),
            ("site","Site","Pesquisar CNPJ no Google","Média"),
            ("decisores","Decisores","Buscar no LinkedIn","Alta")]
    for r in results:
        for campo,label,acao,prio in campos:
            val=r.get(campo)
            empty=not val or val=="não encontrado" or val==[] or val=="—"
            if empty:
                bg=PatternFill("solid",fgColor=lgray if row4%2==0 else white)
                vals=[r.get("cnpj",""),r.get("razao_social",""),label,"Não encontrado nas fontes",acao,prio]
                for j,v in enumerate(vals,1):
                    cell=ws4.cell(row=row4,column=j,value=v)
                    cell.fill=bg; cell.font=Font(name="Arial",size=10)
                row4+=1
    ws4.freeze_panes="A2"
    for i,w in enumerate([20,30,18,30,40,12],1):
        ws4.column_dimensions[get_column_letter(i)].width=w

    # Branding
    for ws in wb.worksheets:
        last=ws.max_row+2 if ws.max_row else 3
        c=ws.cell(row=last,column=1,value="Gerado pelo Enriquece · CredSeller · "+datetime.now().strftime("%d/%m/%Y"))
        c.font=Font(italic=True,color="8896B3",size=9,name="Arial")

    tmp=tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    return FileResponse(tmp,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"enriquece_{sid[:8]}_{filter}.xlsx")

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT",8000)))
