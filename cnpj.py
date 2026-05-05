from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import os

NAVY  = "0D1B3E"
BLUE  = "1E4DB7"
GREEN = "00C896"
AMBER = "F59E0B"
RED   = "EF4444"
WHITE = "FFFFFF"
LGRAY = "F4F6FB"
MGRAY = "C8D0E0"

def _hdr(ws, row, cols, fill_hex=NAVY, font_hex=WHITE, sz=11):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_hex, size=sz, name="Arial")
    al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill = fill; c.font = font; c.alignment = al

def _row(ws, row_num, values, odd=True):
    bg = PatternFill("solid", fgColor=LGRAY) if odd else PatternFill("solid", fgColor=WHITE)
    al = Alignment(vertical="center", wrap_text=False)
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=v if v != "não encontrado" else "—")
        c.fill = bg; c.alignment = al; c.font = Font(size=10, name="Arial")

def _score_cell(ws, row_num, col_num, score):
    colors = {"alta": ("E6F9F4","00A87B"), "media": ("FFF8E6","B45309"), "baixa": ("FEF2F2","DC2626")}
    labels = {"alta": "Alta", "media": "Média", "baixa": "Baixa"}
    bg, fg = colors.get(score, (LGRAY, "333333"))
    c = ws.cell(row=row_num, column=col_num, value=labels.get(score, score))
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=fg, size=10, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")

def _freeze_filter(ws, up_to_col):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(up_to_col)}{ws.max_row or 1}"

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel(session: dict, out_path: str):
    results     = session.get("results", [])
    invalid_rows= session.get("invalid_rows", [])
    duplicates  = [r for lot in session.get("lots",[]) for r in lot if r.get("reason") == "Duplicado"]

    wb = Workbook()
    wb.remove(wb.active)

    # ── ABA 1: RESUMO ──────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumo Executivo")
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 40
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "ENRIQUECE — Resumo do Processamento"
    c.font  = Font(bold=True, color=WHITE, size=14, name="Arial")
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    total    = len(results)
    completo = sum(1 for r in results if r.get("status") == "completo")
    sem_dados= sum(1 for r in results if r.get("status") == "sem dados")
    c_tel    = sum(1 for r in results if r.get("telefone_cad","") not in ("","não encontrado","—"))
    c_email  = sum(1 for r in results if r.get("email_cad","") not in ("","não encontrado","—")
                   or r.get("email_comercial","") not in ("","não encontrado","—"))
    c_site   = sum(1 for r in results if r.get("site","") not in ("","não encontrado","—"))
    c_dec    = sum(1 for r in results if r.get("decisores"))
    scores   = [r.get("score_geral","") for r in results]
    avg_sc   = "Alta" if scores.count("alta") > total*0.6 else "Média" if scores.count("alta") > total*0.3 else "Baixa"

    kv = [
        ("Total de CNPJs processados", total),
        ("Empresas completas", completo),
        ("Empresas sem dados suficientes", sem_dados),
        ("Telefones encontrados", c_tel),
        ("E-mails encontrados", c_email),
        ("Sites encontrados", c_site),
        ("Decisores encontrados", c_dec),
        ("Score médio de confiança", avg_sc),
        ("Data do processamento", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Gerado por", "Enriquece — CredSeller"),
    ]
    for i, (k, v) in enumerate(kv, 3):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True, size=11, name="Arial", color=NAVY)
        vc = ws1.cell(row=i, column=2, value=v)
        vc.font = Font(size=11, name="Arial")
        if i % 2 == 0:
            for col in range(1, 3):
                ws1.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LGRAY)

    _set_widths(ws1, [35, 30, 20, 20])

    # ── ABA 2: BASE ENRIQUECIDA ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Base Enriquecida")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 32
    cols2 = ["CNPJ","Razão Social","Nome Fantasia","Situação","Data Abertura",
             "CNAE Principal","Porte","Cidade","Estado","CEP","Endereço",
             "Site Oficial","LinkedIn Empresa","Telefone","WhatsApp",
             "E-mail Geral","E-mail Comercial","E-mail Financeiro","E-mail Parcerias",
             "Status","Score Geral","Observações"]
    _hdr(ws2, 1, cols2)
    for idx, r in enumerate(results, 2):
        vals = [
            r.get("cnpj",""), r.get("razao_social",""), r.get("nome_fantasia",""),
            r.get("situacao",""), r.get("data_abertura",""), r.get("cnae_principal",""),
            r.get("porte",""), r.get("cidade",""), r.get("estado",""),
            r.get("cep",""), r.get("endereco",""), r.get("site",""),
            r.get("linkedin_empresa",""), r.get("telefone_cad",""),
            r.get("whatsapp",""), r.get("email_cad",""),
            r.get("email_comercial",""), r.get("email_financeiro",""),
            r.get("email_parcerias",""), r.get("status",""),
            None, r.get("observacoes",""),
        ]
        _row(ws2, idx, vals, idx % 2 == 0)
        _score_cell(ws2, idx, 21, r.get("score_geral","baixa"))
    _freeze_filter(ws2, len(cols2))
    _set_widths(ws2, [18,30,25,14,14,35,12,18,8,12,35,30,30,16,16,28,28,28,28,12,12,30])

    # ── ABA 3: DECISORES ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Decisores")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 32
    cols3 = ["CNPJ","Empresa","Nome do Decisor","Cargo","Área",
             "LinkedIn","E-mail","Telefone","Prioridade","Por que abordar","Score","Fonte"]
    _hdr(ws3, 1, cols3, fill_hex=BLUE)
    row3 = 2
    for r in results:
        for d in r.get("decisores", []):
            vals = [r.get("cnpj",""), r.get("razao_social",""),
                    d.get("nome",""), d.get("cargo",""), d.get("area",""),
                    d.get("linkedin",""), d.get("email",""), d.get("phone",""),
                    d.get("prioridade",""), d.get("motivo",""),
                    None, d.get("fonte","")]
            _row(ws3, row3, vals, row3 % 2 == 0)
            _score_cell(ws3, row3, 11, d.get("score","baixa"))
            # Prioridade color
            prio = d.get("prioridade","")
            pc = ws3.cell(row=row3, column=9)
            if prio == "Alta":
                pc.fill = PatternFill("solid", fgColor="E6F9F4")
                pc.font = Font(bold=True, color="00A87B", size=10, name="Arial")
            elif prio == "Média":
                pc.fill = PatternFill("solid", fgColor="FFF8E6")
                pc.font = Font(bold=True, color="B45309", size=10, name="Arial")
            row3 += 1
    _freeze_filter(ws3, len(cols3))
    _set_widths(ws3, [18,28,24,28,18,35,30,16,12,45,12,14])

    # ── ABA 4: FONTES ─────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Fontes")
    ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 32
    cols4 = ["CNPJ","Empresa","Dado Encontrado","Fonte","URL","Tipo","Confiança","Data"]
    _hdr(ws4, 1, cols4, fill_hex="1A2F5E")
    row4 = 2
    for r in results:
        for f in r.get("fontes", []):
            vals = [r.get("cnpj",""), r.get("razao_social",""),
                    f.get("dado",""), f.get("fonte",""), f.get("url",""),
                    f.get("tipo",""), None, f.get("data","")]
            _row(ws4, row4, vals, row4 % 2 == 0)
            _score_cell(ws4, row4, 7, f.get("confianca","baixa"))
            row4 += 1
    _freeze_filter(ws4, len(cols4))
    _set_widths(ws4, [18,28,22,16,45,20,12,14])

    # ── ABA 5: PENDÊNCIAS ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Pendências")
    ws5.sheet_view.showGridLines = False
    ws5.row_dimensions[1].height = 32
    cols5 = ["CNPJ","Empresa","Campo Pendente","Motivo","Próxima Ação","Prioridade"]
    _hdr(ws5, 1, cols5, fill_hex=AMBER, font_hex=NAVY)
    row5 = 2
    campos_check = [
        ("telefone_cad","Telefone","Buscar no Google Maps / site oficial","Alta"),
        ("email_cad","E-mail","Buscar via Hunter.io ou site","Alta"),
        ("site","Site oficial","Pesquisar CNPJ + nome no Google","Média"),
        ("linkedin_empresa","LinkedIn empresa","Buscar por nome fantasia no LinkedIn","Média"),
        ("decisores","Decisores","Buscar executivos no LinkedIn Sales Navigator","Alta"),
    ]
    for r in results:
        for campo, label, acao, prio in campos_check:
            val = r.get(campo)
            empty = not val or val == "não encontrado" or val == [] or val == "—"
            if empty:
                vals = [r.get("cnpj",""), r.get("razao_social",""),
                        label, "Não encontrado nas fontes disponíveis", acao, prio]
                _row(ws5, row5, vals, row5 % 2 == 0)
                row5 += 1
    _freeze_filter(ws5, len(cols5))
    _set_widths(ws5, [18,28,18,35,45,14])

    # ── ABA 6: INVÁLIDOS ──────────────────────────────────────────────────────
    ws6 = wb.create_sheet("CNPJs Inválidos")
    ws6.sheet_view.showGridLines = False
    ws6.row_dimensions[1].height = 32
    cols6 = ["CNPJ Original","Motivo","Linha Original","Status"]
    _hdr(ws6, 1, cols6, fill_hex=RED)
    for idx, r in enumerate(invalid_rows, 2):
        vals = [r.get("raw",""), r.get("reason",""), r.get("row",""), "Inválido/Duplicado"]
        _row(ws6, idx, vals, idx % 2 == 0)
    _freeze_filter(ws6, len(cols6))
    _set_widths(ws6, [22,40,16,20])

    # Branding note
    for ws in wb.worksheets:
        last = ws.max_row + 2 if ws.max_row else 3
        c = ws.cell(row=last, column=1, value="Gerado pelo Enriquece · CredSeller · " + datetime.now().strftime("%d/%m/%Y"))
        c.font = Font(italic=True, color="8896B3", size=9, name="Arial")

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
